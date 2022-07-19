import selenium
from selenium import webdriver
import time
import art
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains


ART = art.text2art('EGE CHECKER')
print(ART)
d = {}
kubgu = []
techgu = []
techgu_balls = []
solved_peoples = []
browser = webdriver.Chrome("C:\CH\chromedriver.exe")
browser.set_window_size(1200, 720)


def parsekubgu():
    browser.get('http://ftp.kubsu.ru/alpha/09.03.03_793_ofo_b.html')
    names = browser.find_elements(by=By.CLASS_NAME, value='style_14')
    first = names[0].text
    kubgu.append(first)
    for i in range(len(names)-1):
        kubgu.append(names[i+1].text)
    print('Подано в КУБГУ: ', len(names))
    print(kubgu)


def parsetechgu():
    browser.get('https://ent.kubstu.ru/')
    time.sleep(3)
    names1 = browser.find_elements(by=By.CSS_SELECTOR, value='td:nth-child(1)')
    print(names1)
    balls1 = browser.find_elements(by=By.CSS_SELECTOR, value='td.sum')
    print('Подано в КУБГТУ: ', len(names1))
    for i in range(len(names1)-1):
        techgu.append(names1[i + 1].text)
        techgu_balls.append(balls1[i+1].text)
        d[names1[i + 1].text] = balls1[i+1].text
    print(techgu_balls)
    print(techgu)
    print(d)


parsekubgu()
parsetechgu()
maxpeoples = 0
minpeoples = 0
ballspeople = []
k=1

if len(kubgu) > len(techgu):
    maxpeoples = len(kubgu)
    print('Больше людей в КУБГУ', maxpeoples)
    minpeoples = len(techgu)
else:
    maxpeoples = len(techgu)
    print('Больше людей в ТЕХГУ', maxpeoples)
    minpeoples = len(kubgu)

for j in range(maxpeoples-1):
    x = techgu[j+1]
    if x in kubgu:
        k = k+1
        solved_peoples.append(x) # записываем фИО человека
        excel_file = openpyxl.load_workbook('refund_id.xlsx')
        excel_sheet = excel_file['Лист1']
        excel_sheet.cell(row=k, column=1).value = techgu[j+1]
        excel_sheet.cell(row=k, column=2).value = techgu_balls[j+1]
        excel_file.save('refund_id.xlsx')


print('----- Ниже указан список совпавших людей -----')
print(solved_peoples)
print('Всего совпало', len(solved_peoples), 'человек')
with open("файл.txt", "w", encoding="utf-8") as file:
    for item in solved_peoples:
        print(item, file=file)
print('Данные об учениках записаны в файл')